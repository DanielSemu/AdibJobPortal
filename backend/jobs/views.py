from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.exceptions import ValidationError
from rest_framework.permissions import IsAuthenticated,AllowAny
from rest_framework.exceptions import NotFound
from django.shortcuts import get_object_or_404
from rest_framework.decorators import api_view, permission_classes
from .serializers import ApplicantSerializer,ContactUsSerializer ,JobCategorySerializer,JobSerializer
from .models import Applicant, Job,ContactUs,JobCategory,JobDetail,TempApplicant,Criteria
from django.utils import timezone
from django.utils.timezone import now
import chardet
from io import BytesIO, TextIOWrapper
import csv
from dateutil.parser import parse
from authApi.permissions import IsAdminRole,ViewJobRole,IsHrCheckerRole
# sdfdsfds
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
import io

    
@api_view(['GET'])
def generate_applicants_pdf(request):
    job_id = request.GET.get("job_id")

    if job_id:
        applicants = Applicant.objects.filter(status="Accepted", job_id=job_id)
        print(applicants)
    else:
        applicants = Applicant.objects.filter(status="Accepted")
    job=Job.objects.get(id=job_id)
    template = get_template("applicants_report.html")
    html = template.render({"applicants": applicants,"job":job})

    result = io.BytesIO()
    pdf = pisa.pisaDocument(io.BytesIO(html.encode("UTF-8")), result)

    if not pdf.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')
    return HttpResponse("PDF generation failed", status=500)

class ExportEmployeeDataView(APIView):
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Applicant.objects.filter(status="Accepted")

    def get(self, request):
        applications = self.get_queryset()
        if not applications.exists():
            return Response({"message": "No applications found."}, status=status.HTTP_404_NOT_FOUND)

        serializer = ApplicantSerializer(applications, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request):
        applications = self.get_queryset()
        if not applications.exists():
            return Response({"message": "No applications found."}, status=status.HTTP_404_NOT_FOUND)

        serializer = ApplicantSerializer(applications, many=True)
        employees = serializer.data

        wb = Workbook()
        ws = wb.active
        ws.title = "Employees"

        headers = [
            "No", "Full Name", "Gender", "Birth Date", "Applied For", "Work Place", "Status",
            "Edu-Organization", "Edu-Level", "Filed-Study",
            "Job-Position", "Organization", "From", "To", "Banking Experience",
            "Certificate-Title", "Company"
        ]
        ws.append(headers)

        for col in range(1, len(headers) + 1):
            ws.cell(row=1, column=col).font = Font(bold=True)
            ws.cell(row=1, column=col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        row_index = 2
        for i, emp in enumerate(employees, start=1):
            educations = emp.get("educations", [])
            experiences = emp.get("experiences", [])
            certifications = emp.get("certifications", [])

            max_len = max(len(educations), len(experiences), len(certifications))

            for j in range(max_len):
                edu = educations[j] if j < len(educations) else {}
                exp = experiences[j] if j < len(experiences) else {}
                cert = certifications[j] if j < len(certifications) else {}

                ws.cell(row_index, 8, edu.get("education_organization", ""))
                ws.cell(row_index, 9, edu.get("education_level", ""))
                ws.cell(row_index, 10, edu.get("field_of_study", ""))

                ws.cell(row_index, 11, exp.get("job_title", ""))
                ws.cell(row_index, 12, exp.get("company_name", ""))
                ws.cell(row_index, 13, exp.get("from_date", ""))
                ws.cell(row_index, 14, exp.get("to_date", ""))
                ws.cell(row_index, 15, exp.get("banking_experience", ""))

                ws.cell(row_index, 16, cert.get("certificate_title", ""))
                ws.cell(row_index, 17, cert.get("awarding_company", ""))

                if j == 0:
                    ws.cell(row_index, 1, i)
                    ws.cell(row_index, 2, emp["full_name"])
                    ws.cell(row_index, 3, emp["gender"])
                    ws.cell(row_index, 4, emp["birth_date"])
                    ws.cell(row_index, 5, emp["job_name"])
                    ws.cell(row_index, 6, emp["selected_work_place"])
                    ws.cell(row_index, 7, emp["status"])

                row_index += 1

            for col in [1, 2, 3, 4, 5, 6, 7]:
                ws.merge_cells(start_row=row_index - max_len, start_column=col, end_row=row_index - 1, end_column=col)
                ws.cell(row=row_index - max_len, column=col).alignment = Alignment(vertical="top", wrap_text=True)

        column_widths = [5, 25, 15, 15, 20, 20, 20, 25, 25, 25, 12, 20, 15, 15, 12, 20, 20]
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = 'attachment; filename="employee_export.xlsx"'
        wb.save(response)

        return response

class AdminJobView(APIView):
    
    permission_classes =[IsAuthenticated,ViewJobRole]
    
    def get(self, request, id=None, *args, **kwargs):
        if id:
            job=get_object_or_404(Job, id=id)
            serializer = JobSerializer(job)
        else:
            jobs=Job.objects.all().order_by('-updated_at')
            serializer = JobSerializer(jobs, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    def post(self, request, *args, **kwargs):
        serializer = JobSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def put(self, request, id=None, *args, **kwargs):
        job = get_object_or_404(Job, id=id)
        serializer = JobSerializer(job, data=request.data, partial=True)
        
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        
        # print("Validation errors:", serializer.errors)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, id=None, *args, **kwargs):
        job = get_object_or_404(Job, id=id)
        job.delete()
        return Response({"message": "Job deleted successfully."}, status=status.HTTP_204_NO_CONTENT)

class JobView(APIView):
    def get(self, request, id=None, *args, **kwargs):
        if id:
            job = get_object_or_404(Job, id=id)  # Get a single job
            serializer = JobSerializer(job)
        else:
            today = timezone.now().date()
            jobs = Job.objects.filter(
            status="Active",
            post_date__lte=today,
            application_deadline__gte=today
            )
            serializer = JobSerializer(jobs, many=True)
        
        return Response(serializer.data, status=status.HTTP_200_OK)
class ExpiredJobView(APIView):
    permission_classes =[IsAuthenticated,ViewJobRole]
    def get(self, request, id=None, *args, **kwargs):
        # jobs = Job.objects.filter(status="Active",application_deadline__lt=timezone.now().date())
        jobs = Job.objects.filter(status="Active")

        serializer = JobSerializer(jobs, many=True)
        
        return Response(serializer.data, status=status.HTTP_200_OK)
    
    
class JobBulkUploadView(APIView):
    permission_classes = [IsAuthenticated, ViewJobRole]

    def post(self, request, *args, **kwargs):
        csv_file = request.FILES.get('file')
        if not csv_file:
            return Response({"error": "No file provided"}, status=status.HTTP_400_BAD_REQUEST)

        # Detect encoding from sample
        sample = csv_file.read(1024)
        encoding = chardet.detect(sample)['encoding'] or 'utf-8'
        csv_file.seek(0)  # rewind

        decoded_file = TextIOWrapper(csv_file.file, encoding=encoding)
        reader = csv.DictReader(decoded_file)

        jobs_to_create = []
        errors = []
        row_number = 1  # for clearer error messages

        for row in reader:
            try:
                row_number += 1

                # Check for required fields
                required_fields = ['title', 'category', 'location', 'job_type', 'description', 'application_deadline', 'post_date']
                for field in required_fields:
                    if not row.get(field):
                        raise ValueError(f"Missing required field '{field}'.")

                try:
                    category_obj = JobCategory.objects.get(name=row['category'])
                except JobCategory.DoesNotExist:
                    raise ValueError(f"Category '{row['category']}' not found.")

                try:
                    parsed_deadline = parse(row['application_deadline'], dayfirst=False).date()
                    parsed_post_date = parse(row['post_date'], dayfirst=False).date()
                except Exception:
                    raise ValueError("Invalid date format in 'application_deadline' or 'post_date'.")

                job = Job(
                    vacancy_number=row.get('vacancy_number') or None,
                    title=row['title'],
                    job_grade=row.get('job_grade') or None,
                    company=row.get('company') or "Addis Bank S.C",
                    category=category_obj,
                    location=row['location'],
                    job_type=row['job_type'],
                    salary=row.get('salary') or "As per Companies Salary Scale",
                    description=row['description'],
                    post_date=parsed_post_date,
                    application_deadline=parsed_deadline,
                    show_experience=str(row.get('show_experience', 'True')).lower() == 'true',
                    status=row.get('status') or "InActive",
                    created_at=now(),
                    updated_at=now()
                )

                jobs_to_create.append(job)

            except Exception as e:
                errors.append({"row": row_number, "error": str(e)})

        if errors:
            return Response(
                {"message": "Some rows could not be processed.", "errors": errors},
                status=status.HTTP_400_BAD_REQUEST
            )

        Job.objects.bulk_create(jobs_to_create)
        return Response({"message": f"{len(jobs_to_create)} jobs uploaded successfully!"}, status=status.HTTP_201_CREATED)
class JobDetailBulkUploadView(APIView):
    permission_classes =[IsAuthenticated,ViewJobRole]
    def post(self, request, job_id, *args, **kwargs):
        csv_file = request.FILES.get('file')
        if not csv_file:
            return Response({"error": "No file provided"}, status=status.HTTP_400_BAD_REQUEST)

        # Read small sample for encoding detection
        sample = csv_file.read(1024)
        encoding = chardet.detect(sample)['encoding'] or 'utf-8'

        # Go back to beginning of file
        csv_file.seek(0)

        decoded_file = TextIOWrapper(csv_file.file, encoding=encoding)
        reader = csv.DictReader(decoded_file)

        # 🔵 Get the Job once
        try:
            job = Job.objects.get(id=job_id)
        except Job.DoesNotExist:
            return Response({"error": f"Job with id '{job_id}' not found."},
                            status=status.HTTP_400_BAD_REQUEST)

        details_to_create = []

        for row in reader:
            detail = JobDetail(
                job=job,
                detail_type=row['detail_type'],
                description=row['description'],
            )
            details_to_create.append(detail)

        JobDetail.objects.bulk_create(details_to_create)

        return Response({"message": "Job Details uploaded successfully!"}, status=status.HTTP_201_CREATED)


class JobCategoryView(APIView):
    # permission_classes =[IsAuthenticated,ViewJobRole]
    def get_permissions(self):
        if self.request.method in ['POST', 'PUT', 'DELETE']:
            return [IsAuthenticated(), ViewJobRole()]  # Auth + Role check
        return [AllowAny()]  # <-- 🛠 Allow GET for everyone
    def get(self, request, id=None, *args, **kwargs):
        if id:
            category = get_object_or_404(JobCategory, id=id)  # Get a single category
            serializer = JobCategorySerializer(category)
        else:
            categories = JobCategory.objects.all()  # Get all categories
            serializer = JobCategorySerializer(categories, many=True)
        
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request, *args, **kwargs):
        serializer = JobCategorySerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def put(self, request, id=None, *args, **kwargs):
        category = get_object_or_404(JobCategory, id=id)
        serializer = JobCategorySerializer(category, data=request.data, partial=True)
        
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, id=None, *args, **kwargs):
        category = get_object_or_404(JobCategory, id=id)
        category.delete()
        return Response({"message": "Job category deleted successfully."}, status=status.HTTP_204_NO_CONTENT)



class ApplicantAPIView(APIView):
    def post(self, request):
        # print(request.data)
        serializer = ApplicantSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(
                {"status": "success", "message": "Application submitted successfully"},
                status=status.HTTP_201_CREATED
            )
        print("Validation Errors:", serializer.errors)
        return Response(
            {"status": "error", "errors": serializer.errors},
            status=status.HTTP_400_BAD_REQUEST
        )
    def get(self, request, id=None, *args, **kwargs):
        if id:
            applicants = Applicant.objects.get(id=id,status="Under Review")
            serializer = ApplicantSerializer(applicants)
        else:
            applicants = Applicant.objects.filter(status="Under Review")
            serializer = ApplicantSerializer(applicants, many=True)
            
        return Response(serializer.data, status=status.HTTP_200_OK)

    def put(self,request, id=None, *args, **kwargs):
        if not id:
            return Response({"error":'Applicant id is required.'},status=status.HTTP_400_BAD_REQUEST)
        
        try:
            applicant=Applicant.objects.get(id=id)
        except Applicant.DoesNotExist:
            return Response({'error': 'Applicant not found.'}, status=status.HTTP_404_NOT_FOUND)
        
        new_status=request.data.get('status')
        if not new_status:
            return Response({'error':"status field is required"}, status=status.HTTP_400_BAD_REQUEST)

        applicant.status=new_status
        applicant.save()
        
        return Response({
            'message':'Status updated successfully',
            'applicant_id':applicant.id,
            'new_status':applicant.status,
        }, status=status.HTTP_200_OK)
      
    
@api_view(['GET'])
@permission_classes([IsAuthenticated])    
def getUserApplications(request):
    user_email=request.user.email
    applications=Applicant.objects.filter(email=user_email)
    
    if not applications.exists():
        return Response({"message": "No applications found for this user."}, status=status.HTTP_404_NOT_FOUND)
    serializer=ApplicantSerializer(applications, many=True)
    return  Response(serializer.data, status=status.HTTP_200_OK)

@api_view(['GET'])
@permission_classes([IsAuthenticated])  
def getUnderReviewApplicants(request):
    applicants=Applicant.objects.filter(status='Under Review')
    serializer=ApplicantSerializer(applicants, many=True)
    return Response(serializer.data, status=status.HTTP_200_OK)

class ActiveJobApplicants(APIView):
    def get(self, request, id=None, *args, **kwargs):
        if id:
            applicants = Applicant.objects.filter(job_id=id)
            serializer = ApplicantSerializer(applicants, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
        

class FilterApplicantsView(APIView):
    def post(self,request):
        data=request.data
        
        selected_job=data.get('selectedJob')
        selected_location=data.get('selectedLocation')
        min_experience_years =data.get('minExperienceYears')
        gender=data.get('gender')
        min_cgpa=data.get('minCGPA')
        min_exit=data.get('minExit')
        
        applicants=Applicant.objects.filter(status = 'Pending')
        
        if selected_job:
            applicants=applicants.filter(job__id=selected_job)
        
        if selected_location:
            # applicants = applicants.filter(job__location__icontains=selected_location)
            applicants = applicants.filter(selected_work_place__icontains=selected_location)
            
        if gender:
            applicants=applicants.filter(gender=gender)
        
        filtered_applicants=[]
        for applicant in applicants:
            total_experience_years=0
            for experience in applicant.experiences.all():
                if experience.from_date and experience.to_date:
                    duration=(experience.to_date - experience.from_date).days / 365
                    total_experience_years += duration
            
            
            if min_experience_years  and total_experience_years < float(min_experience_years ):
                continue
            
            educations= applicant.educations.all()
            
            if not educations.exists():
                continue
            
            highest_education=educations.order_by('-graduation_year').first()
            if min_cgpa and float(highest_education.cgpa) < float(min_cgpa):
                continue
            
            if min_exit and float(highest_education.exit_exam) < float(min_exit):
                continue  # Skip if Exit exam score is low
            # applicant.status="Under Review"
            # applicant.save()
            filtered_applicants.append(applicant)
        serializer = ApplicantSerializer(filtered_applicants, many=True)

        return Response(serializer.data, status=status.HTTP_200_OK)

class ConfirmFilteredApplicants(APIView):
    def post(self, request):
        data = request.data
        confirm = data.get('confirm')
        applicant_ids = data.get('applicant_ids', [])
        criteria_data = data.get('criteria', {})
        print(confirm,applicant_ids,criteria_data)
        if not confirm or not applicant_ids:
            return Response({'error': 'Invalid request'}, status=status.HTTP_400_BAD_REQUEST)

        # Update applicants
        updated_count = Applicant.objects.filter(id__in=applicant_ids).update(status='Under Review')

        # Save criteria (you can customize your model)
        Criteria.objects.create(
            job_id=criteria_data.get('selectedJob'),
            location=criteria_data.get('selectedLocation'),
            min_experience_years=criteria_data.get('minExperienceYears'),
            gender=criteria_data.get('gender'),
            min_cgpa=criteria_data.get('minCGPA'),
            min_exit_score=criteria_data.get('minExit'),
            matched_applicants=updated_count,
            timestamp=timezone.now()
        )

        return Response({'message': f'{updated_count} applicants updated and criteria saved.'}, status=status.HTTP_200_OK)


class ContactUsAPIView(APIView):  
    # ✅ GET: Retrieve all contact messages  
    def get(self, request, *args, **kwargs):  
        contacts = ContactUs.objects.all()  
        serializer = ContactUsSerializer(contacts, many=True)  
        return Response(serializer.data, status=status.HTTP_200_OK)  

    # ✅ POST: Create a new contact message  
    def post(self, request, *args, **kwargs):  
        print(request.data)
        serializer = ContactUsSerializer(data=request.data)  
        if serializer.is_valid():  
            serializer.save()  
            return Response(serializer.data, status=status.HTTP_201_CREATED)  
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)  

    # ✅ PUT: Update an existing contact message  
    def put(self, request, *args, **kwargs):  
        contact_id = kwargs.get('pk')  
        try:  
            contact = ContactUs.objects.get(pk=contact_id)  
        except ContactUs.DoesNotExist:  
            return Response({"error": "Contact not found"}, status=status.HTTP_404_NOT_FOUND)  

        serializer = ContactUsSerializer(contact, data=request.data, partial=True)  
        if serializer.is_valid():  
            serializer.save()  
            return Response(serializer.data, status=status.HTTP_200_OK)  
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)  

    # ✅ DELETE: Remove a contact message  
    def delete(self, request, *args, **kwargs):  
        contact_id = kwargs.get('pk')  
        try:  
            contact = ContactUs.objects.get(pk=contact_id)  
            contact.delete()  
            return Response({"message": "Deleted successfully"}, status=status.HTTP_204_NO_CONTENT)  
        except ContactUs.DoesNotExist:  
            return Response({"error": "Contact not found"}, status=status.HTTP_404_NOT_FOUND)  

