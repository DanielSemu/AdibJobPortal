
# ====================
# Standard Library
# ====================
import io
import urllib.parse

# ====================
# Third-Party Libraries
# ====================
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from xhtml2pdf import pisa
 
# ====================
# Django Core Imports
# ====================
from django.http import HttpResponse
from django.template.loader import get_template
from django.utils import timezone
from django.shortcuts import get_object_or_404

# ====================
# Django REST Framework
# ====================

from rest_framework import status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import api_view, permission_classes
from rest_framework.parsers import JSONParser, MultiPartParser


# ====================
# Local App Imports
# ====================
from .serializers import (
    ApplicantSerializer,
    CriteriaSerializer
)

from .models import (
    Applicant,
    Job,
    Criteria
)
from authApi.permissions import (
    IsAdminRole,
    ViewJobRole,
    IsHrCheckerRole,
)
from authApi.backends import ApplicantJWTAuthentication




#========================
#User Apply For a Job
#=========================
class UserApplyForJobAPIView(APIView):
    parser_classes = [JSONParser, MultiPartParser]
    def post(self, request):
        # print(request.data)
        serializer = ApplicantSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(
                {"status": "success", "message": "Application submitted successfully"},
                status=status.HTTP_201_CREATED
            )
        print("Validation Errors:", serializer.errors)
        return Response(
            {"status": "error", "errors": serializer.errors},
            status=status.HTTP_400_BAD_REQUEST
        )
    

#============================#
#=====User's Applcations=====#

class UserApplicationAPIView(APIView):
    authentication_classes = [ApplicantJWTAuthentication]  # 🔥 critical!
    permission_classes = [IsAuthenticated]

    def get(self, request, id=None, *args, **kwargs):
        user = request.user  # now this will be the correct ApplicantUser

        if id:
            applicant = get_object_or_404(Applicant, id=id, email=user.email)
            serializer = ApplicantSerializer(applicant)
        else:
            applicants = Applicant.objects.filter(email=user.email)
            serializer = ApplicantSerializer(applicants, many=True)

        return Response(serializer.data, status=status.HTTP_200_OK)




#========================
# Admin Applications View
#========================
class AdminApplicationsAPIView(APIView):
    permission_classes=[IsAuthenticated , ViewJobRole]
    
    def get(self, request, id=None, *args, **kwargs):
        if id:
            # This Id job Id 
            applicants = Applicant.objects.get(id=id)
            serializer = ApplicantSerializer(applicants)
        else:
            applicants=Applicant.objects.all()
            status_param=request.query_params.get("status")
            if status_param:
                applicants=applicants.filter(status=status_param)
            #Filter By Criteria
            criteria_selected=request.query_params.get("criteria")
            if criteria_selected:
                print(criteria_selected)
                applicants=applicants.filter(selectedCriteria=criteria_selected)
            
            serializer=ApplicantSerializer(applicants, many=True)
 
        return Response(serializer.data, status=status.HTTP_200_OK)
    

    def put(self,request, id=None, *args, **kwargs):
        if not id:
            return Response({"error":'Applicant id is required.'},status=status.HTTP_400_BAD_REQUEST)
        
        try:
            applicant=Applicant.objects.get(id=id)
        except Applicant.DoesNotExist:
            return Response({'error': 'Applicant not found.'}, status=status.HTTP_404_NOT_FOUND)
        
        new_status=request.data.get('status')
        if not new_status:
            return Response({'error':"status field is required"}, status=status.HTTP_400_BAD_REQUEST)

        applicant.status=new_status
        applicant.save()
        
        return Response({
            'message':'Status updated successfully',
            'applicant_id':applicant.id,
            'new_status':applicant.status,
        }, status=status.HTTP_200_OK)


class AdminRemoveApplicant(APIView):
    permission_classes=[IsAuthenticated, ViewJobRole]
    
    def put(self, request, id=None, *args, **kwargs):
        try:
            applicant=get_object_or_404(Applicant, id=id)
            remark=request.data.get("remark")
           
            applicant.status="Rejected"
            applicant.remark=remark
            applicant.save()
            return Response({"message": "Applicant updated successfully."}, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

#===================================
#Admin Fetch Applicants Based on Jobs
#====================================
class AdminFetchJobApplicants(APIView):
    permission_classes=[IsAuthenticated, ViewJobRole]
    
    def get(self,request,id=None,*args, **kwargs):
        if id:
            applicants = Applicant.objects.filter(job_id=id)
            #Filter By Applicant Status
            status_param = request.query_params.get("status")
            if status_param:
                applicants=applicants.filter(status=status_param)
                
            #Filter By WorkPlace Selected
            workPlace_param=request.query_params.get("workPlace")
            if workPlace_param:
                applicants=applicants.filter(selected_work_place=workPlace_param)
            
            
            serializer=ApplicantSerializer(applicants, many=True)
            
        return Response(serializer.data, status=status.HTTP_200_OK)
            
            


#==========================
# Admin Filter Applications
#==========================
class FilterApplicantsView(APIView):
    permission_classes = [IsAuthenticated, ViewJobRole]

    def post(self, request):
        data = request.data

        selected_job = data.get('selectedJob')
        selected_location = data.get('selectedLocation')
        min_experience_years = data.get('minExperienceYears')
        max_experience_years = data.get('maxExperienceYears')
        gender = data.get('gender')
        min_cgpa = data.get('minCGPA')
        min_exit = data.get('minExit')
        program = data.get('program')  # new
        institution = data.get('institution')  # new
        graduation_year = data.get('graduationYear')  # optional

        applicants = Applicant.objects.filter(job__id=selected_job, status='Pending')

        if selected_location:
            applicants = applicants.filter(selected_work_place=selected_location)

        if gender:
            applicants = applicants.filter(gender=gender)

        filtered_applicants = []
        for applicant in applicants:
            selected_education = applicant.educations.filter(user_for_application=True).first()
            if not selected_education:
                
                continue

            # 1. Min CGPA
            if min_cgpa and float(selected_education.cgpa) < float(min_cgpa):
                continue

            # 2. Min Exit Exam
            if min_exit and float(selected_education.exit_exam) < float(min_exit):
                continue

            # 3. Program
            if program and selected_education.program.lower() != program.lower():
                continue

            # 4. Institution
            if institution and selected_education.institution.lower() != institution.lower():
                continue


            # 5. Graduation Year
            if graduation_year and str(selected_education.graduation_year)[:4] != str(graduation_year)[:4]:
                continue

            # 6. Experience check
            total_experience_years = 0
            for exp in applicant.experiences.all():
                if exp.from_date and exp.to_date:
                    duration = (exp.to_date - exp.from_date).days / 365
                    total_experience_years += duration
            if min_experience_years or max_experience_years:
                min_years = float(min_experience_years) if min_experience_years else 0
                max_years = float(max_experience_years) if max_experience_years else float('inf')
                
                if not (min_years <= total_experience_years <= max_years):
                    continue


            filtered_applicants.append(applicant)

        serializer = ApplicantSerializer(filtered_applicants, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class ConfirmFilteredApplicants(APIView):
    permission_classes = [IsAuthenticated, ViewJobRole]

    def post(self, request):
        data = request.data

        confirm = data.get('confirm')
        applicant_ids = data.get('applicant_ids', [])
        criteria_data = data.get('criteria', {})

        print("DEBUG:", confirm, applicant_ids, criteria_data)

        # Validate basic structure
        if not confirm or not applicant_ids or not isinstance(criteria_data, dict):
            return Response({'error': 'Invalid request'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            criteria = Criteria.objects.create(
                job_id=criteria_data.get('selectedJob'),
                location=criteria_data.get('selectedLocation'),
                min_experience_years=criteria_data.get('minExperienceYears') or None,
                gender=criteria_data.get('gender') or None,
                min_cgpa=criteria_data.get('minCGPA') or None,
                min_exit_score=criteria_data.get('minExit') or None,
                matched_applicants=len(applicant_ids),
                timestamp=timezone.now()
            )
        except Exception as e:
            return Response({'error': f'Failed to create criteria: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        # Update applicants
        try:
            updated_count = Applicant.objects.filter(id__in=applicant_ids).update(
                status='Shortlisted',
                selectedCriteria=criteria
            )
        except Exception as e:
            return Response({'error': f'Failed to update applicants: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response({
            'message': f'{updated_count} applicants updated and criteria saved.'
        }, status=status.HTTP_200_OK)
    # Revert Selected Applicants
    def put(self, request, id=None, *args, **kwargs):
        if not id:
            return Response({"error": 'Criteria id is required.'}, status=status.HTTP_400_BAD_REQUEST)
        criteria=Criteria.objects.get(id=id)
        if criteria.message_sent:
            return Response({"errro":"Applicant with this criteria is already sned sms you cant revert it"}, status=status.HTTP_400_BAD_REQUEST)
        # Revert applicants
        updated_count = Applicant.objects.filter(
            selectedCriteria=id,
            status="Shortlisted"
        ).update(
            status="Pending",
            selectedCriteria=None
        )

        # Delete the criteria if it exists
        Criteria.objects.filter(id=id).delete()

        return Response({'message': f'{updated_count} applicants reverted and criteria deleted.'}, status=status.HTTP_200_OK)

class CriteriasAPIView(APIView):
    permission_classes=[IsAuthenticated, ViewJobRole]    
    
    def get(self,request):
        criterias=Criteria.objects.all().order_by("message_sent")
        serializer=CriteriaSerializer(criterias, many=True)
        
        return Response(serializer.data, status=status.HTTP_200_OK) 
        



# Send SMS for Shortlisted Applicants  
class SendSMSView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        selectedCriteria = request.data.get("selectedCriteria")
        message = request.data.get("message")
        testMessage = request.data.get("testMessage")
        sender = request.user.phone_number

        if not selectedCriteria or not message:
            return Response({"error": "Missing recipient or message"}, status=status.HTTP_400_BAD_REQUEST)

        selectedApplicant = Applicant.objects.filter(selectedCriteria=selectedCriteria, status="Shortlisted")
        encoded_message = urllib.parse.quote(message)
        base_url = (
            "http://192.168.6.27:9501/api?action=sendmessage"
            "&username=Test&password=Adib@123"
            "&messagetype=SMS:TEXT"
            f"&messagedata={encoded_message}"
        )
        results = []

        if testMessage:
            # Send test message to sender
            url = f"{base_url}&recipient={sender}"
            try:
                response = requests.get(url)
                results.append({
                    "recipient": sender,
                    "status": "sent",
                    "provider_response": response.text
                })
            except Exception as e:
                results.append({
                    "recipient": sender,
                    "status": "failed",
                    "error": str(e)
                })
        else:
            # Send SMS to applicants and update status
            for applicant in selectedApplicant:
                number = applicant.phone
                url = f"{base_url}&recipient={number}"
                try:
                    response = requests.get(url)
                    
                    # Update status after successful SMS
                    applicant.status = "SMS_Sent"
                    applicant.save()

                    results.append({
                        "recipient": number,
                        "status": "sent",
                        "provider_response": response.text
                    })
                except Exception as e:
                    results.append({
                        "recipient": number,
                        "status": "failed",
                        "error": str(e)
                    })
            Criteria.objects.filter(id=selectedCriteria).update(message_sent=True)


        return Response({"status": "completed", "results": results}, status=status.HTTP_200_OK)
 
 
 
    
@api_view(['GET'])
@permission_classes([IsAuthenticated]) 
def generate_applicants_pdf(request):
    job_id = request.GET.get("job_id")

    if job_id:
        applicants = Applicant.objects.filter(status="Accepted", job_id=job_id)
        print(applicants)
    else:
        applicants = Applicant.objects.filter(status="Accepted")
    job=Job.objects.get(id=job_id)
    template = get_template("applicants_report.html")
    html = template.render({"applicants": applicants,"job":job})

    result = io.BytesIO()
    pdf = pisa.pisaDocument(io.BytesIO(html.encode("UTF-8")), result)

    if not pdf.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')
    return HttpResponse("PDF generation failed", status=500)



class ApplicantReportAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        status_param = request.query_params.get("status", "Accepted")
        job_param = request.query_params.get("job_id")

        # Filter applicants based on status and job_id
        if job_param:
            applicants = Applicant.objects.filter(status=status_param, job_id=job_param).prefetch_related('experiences')
        else:
            applicants = Applicant.objects.filter(status=status_param).prefetch_related('experiences')

        # Fetch the job if job_param is provided
        job = None
        if job_param:
            try:
                job = Job.objects.get(id=job_param)
            except Job.DoesNotExist:
                return Response({"error": "Job not found."}, status=404)

        # Add calculated experience fields to each applicant
        for applicant in applicants:
            banking_exp = 0
            non_banking_exp = 0

            for exp in applicant.experiences.all():
                from_date = exp.from_date
                to_date = exp.to_date or date.today()

                duration_days = (to_date - from_date).days
                years = round(duration_days / 365.25, 2)

                if exp.banking_experience:
                    banking_exp += years
                else:
                    non_banking_exp += years

            applicant.total_banking_experience = round(banking_exp, 2)
            applicant.total_non_banking_experience = round(non_banking_exp, 2)
            applicant.total_experience = round(banking_exp + non_banking_exp, 2)

        # Render HTML using template
        template = get_template("report.html")
        html = template.render({
            "applicants": applicants,
            "job": job
        })

        # Generate PDF
        result = io.BytesIO()
        pdf = pisa.pisaDocument(io.BytesIO(html.encode("UTF-8")), result)

        if not pdf.err:
            return HttpResponse(result.getvalue(), content_type='application/pdf')

        return Response({"error": "PDF generation failed"}, status=500)
    
    
class ExportEmployeeDataView(APIView):
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Applicant.objects.filter(status="Accepted")

    def get(self, request):
        applications = self.get_queryset()
        if not applications.exists():
            return Response({"message": "No applications found."}, status=status.HTTP_404_NOT_FOUND)

        serializer = ApplicantSerializer(applications, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request):
        applications = self.get_queryset()
        if not applications.exists():
            return Response({"message": "No applications found."}, status=status.HTTP_404_NOT_FOUND)

        serializer = ApplicantSerializer(applications, many=True)
        employees = serializer.data

        wb = Workbook()
        ws = wb.active
        ws.title = "Employees"

        headers = [
            "No", "Full Name", "Gender", "Birth Date", "Applied For", "Work Place", "Status",
            "Edu-Organization", "Edu-Level", "Filed-Study",
            "Job-Position", "Organization", "From", "To", "Banking Experience",
            "Certificate-Title", "Company"
        ]
        ws.append(headers)

        for col in range(1, len(headers) + 1):
            ws.cell(row=1, column=col).font = Font(bold=True)
            ws.cell(row=1, column=col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        row_index = 2
        for i, emp in enumerate(employees, start=1):
            educations = emp.get("educations", [])
            experiences = emp.get("experiences", [])
            certifications = emp.get("certifications", [])

            max_len = max(len(educations), len(experiences), len(certifications))

            for j in range(max_len):
                edu = educations[j] if j < len(educations) else {}
                exp = experiences[j] if j < len(experiences) else {}
                cert = certifications[j] if j < len(certifications) else {}

                ws.cell(row_index, 8, edu.get("education_organization", ""))
                ws.cell(row_index, 9, edu.get("education_level", ""))
                ws.cell(row_index, 10, edu.get("field_of_study", ""))

                ws.cell(row_index, 11, exp.get("job_title", ""))
                ws.cell(row_index, 12, exp.get("company_name", ""))
                ws.cell(row_index, 13, exp.get("from_date", ""))
                ws.cell(row_index, 14, exp.get("to_date", ""))
                ws.cell(row_index, 15, exp.get("banking_experience", ""))

                ws.cell(row_index, 16, cert.get("certificate_title", ""))
                ws.cell(row_index, 17, cert.get("awarding_company", ""))

                if j == 0:
                    ws.cell(row_index, 1, i)
                    ws.cell(row_index, 2, emp["full_name"])
                    ws.cell(row_index, 3, emp["gender"])
                    ws.cell(row_index, 4, emp["birth_date"])
                    ws.cell(row_index, 5, emp["job_name"])
                    ws.cell(row_index, 6, emp["selected_work_place"])
                    ws.cell(row_index, 7, emp["status"])

                row_index += 1

            for col in [1, 2, 3, 4, 5, 6, 7]:
                ws.merge_cells(start_row=row_index - max_len, start_column=col, end_row=row_index - 1, end_column=col)
                ws.cell(row=row_index - max_len, column=col).alignment = Alignment(vertical="top", wrap_text=True)

        column_widths = [5, 25, 15, 15, 20, 20, 20, 25, 25, 25, 12, 20, 15, 15, 12, 20, 20]
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = 'attachment; filename="employee_export.xlsx"'
        wb.save(response)

        return response




      

